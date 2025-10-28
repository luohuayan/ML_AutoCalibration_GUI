import mlcolorimeter as mlcm
from typing import List, Dict
import cv2
import csv
import os
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook


def datetime_str():
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S.%f")[:-3]


def calculate_sph_cyl_coefficinet(
    colorimeter: mlcm.ML_Colorimeter,
    save_path: str,
    file_name: str,
    nd: mlcm.MLFilterEnum,
    xyz: mlcm.MLFilterEnum,
    exposure: mlcm.pyExposureSetting,
    avg_count: int,
    sph_list: List,
    cyl_list: List,
    roi: mlcm.pyCVRect,
):
    mono = colorimeter.ml_bino_manage.ml_get_module_by_id(1)
    pixel_format = mlcm.MLPixelFormat.MLBayerRG12
    ret = mono.ml_set_pixel_format(pixel_format)
    if not ret.success:
        raise RuntimeError("ml_set_pixel_format error")

    ret = mono.ml_move_nd_syn(nd)
    if not ret.success:
        raise RuntimeError("ml_move_nd_syn error")

    ret = mono.ml_move_xyz_syn(xyz)
    if not ret.success:
        raise RuntimeError("ml_move_xyz_syn error")

    ret = ml_mono.ml_set_exposure(exposure)
    if not ret.success:
        raise RuntimeError("ml_set_exposure error")
    
    save_xlsx = save_path + "\\" + file_name
    wb = Workbook()
    wb.save(save_xlsx)
    wb = load_workbook(save_xlsx)
    wb.remove(wb["Sheet"])
    title = str("coefficient")
    ws = wb.create_sheet(title=title)
    line = []
    ws.append(line)
    line = []

    for i in range(avg_count):
        sph_coefficient = []
        cyl_coefficient = []
        last_sph = 0
        last_cyl = 0
        for sph in sph_list:
            rx = mlcm.pyRXCombination(sph=sph, cyl=0, axis=0)
            ret = ml_mono.ml_set_rx_syn(rx)
            if not ret.success:
                raise RuntimeError("ml_set_rx_syn error")

            ret = ml_mono.ml_capture_image_syn()
            if not ret.success:
                raise RuntimeError("ml_capture_image_syn error")

            img = ml_mono.ml_get_image()
            X, Y, Z = cv2.split(img)
            cv2.imwrite(
                save_path + "\\" +
                mlcm.pyRXCombination_to_str(rx) + ".tif", Y
            )
            gray = cv2.mean(
                Y[roi.y:roi.y+roi.height, roi.x:roi.x+roi.width])[0]
            sph_coefficient.append(gray)

            if sph == 0:
                last_sph = gray
        for i in range(len(sph_coefficient)):
            sph_coefficient[i] = format(sph_coefficient[i] / last_sph, ".3f")

        print("sph coefficient: ")
        print(sph_coefficient)
        line = ["sph_coefficient", *sph_coefficient]
        ws.append(line)
        line = []
        ws.append(line)
        line = []

        for cyl in cyl_list:
            rx = mlcm.pyRXCombination(sph=0, cyl=cyl, axis=0)
            ret = ml_mono.ml_set_rx_syn(rx)
            if not ret.success:
                raise RuntimeError("ml_set_rx_syn error")

            ret = ml_mono.ml_capture_image_syn()
            if not ret.success:
                raise RuntimeError("ml_capture_image_syn error")

            img = ml_mono.ml_get_image()
            X, Y, Z = cv2.split(img)
            cv2.imwrite(
                save_path + "\\" +
                mlcm.pyRXCombination_to_str(rx) + ".tif", Y
            )
            gray = cv2.mean(
                Y[roi.y:roi.y+roi.height, roi.x:roi.x+roi.width])[0]
            cyl_coefficient.append(gray)

            if cyl == 0:
                last_cyl = gray
        for i in range(len(cyl_coefficient)):
            cyl_coefficient[i] = format(cyl_coefficient[i] / last_cyl, ".3f")

        print("cyl coefficient: ")
        print(cyl_coefficient)
        line = ["cyl_coefficient", *cyl_coefficient]
        ws.append(line)
        line = []
        ws.append(line)
        line = []

    wb.save(save_xlsx)
    print("calculate sph cyl coefficient finish")


if __name__ == "__main__":
    eye1_path = r"D:\MLColorimeter\config\EYE1"
    path_list = [
        eye1_path,
    ]
    save_path = r"D:\sph_cyl_coefficient"
    if not os.path.exists(save_path):
        os.makedirs(save_path)
    file_name = "sph_cyl_coefficient.xlsx"
    try:
        # create a ML_Colorimeter system instance
        ml_colorimeter = mlcm.ML_Colorimeter()
        # add mono module into ml_colorimeter system, according to path_list create one or more mono module
        ret = ml_colorimeter.ml_add_module(path_list=path_list)
        if not ret.success:
            raise RuntimeError("ml_add_module error")
        # connect all module in the ml_colorimeter system
        ret = ml_colorimeter.ml_connect()
        if not ret.success:
            raise RuntimeError("ml_connect error")

        module_id = 1
        ml_mono = ml_colorimeter.ml_bino_manage.ml_get_module_by_id(module_id)

        nd = mlcm.MLFilterEnum.ND0
        xyz = mlcm.MLFilterEnum.Y
        exposure = mlcm.pyExposureSetting(mlcm.ExposureMode.Fixed, 100)
        avg_count = 10
        sph_list = [-6, -5, -4, -3, -2, -1, 0, 1, 2, 3, 4]
        cyl_list = [-4, -3.5, -3, -2.5, -2, -1.5, -1, -0.5, 0]

        roi = mlcm.pyCVRect(0, 0, 300, 300)

        calculate_sph_cyl_coefficinet(
            ml_colorimeter,
            save_path,
            file_name,
            nd,
            xyz,
            exposure,
            avg_count,
            sph_list,
            cyl_list,
            roi,
        )
    except Exception as e:
        print(e)
