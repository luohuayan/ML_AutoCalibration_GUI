import mlcolorimeter as mlcm
import os
import cv2
from typing import List, Dict
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import time
from datetime import datetime

__all__ = ["capture_dark_heatmap"]


def preprocess_image(img, block_size):
    height, width = img.shape
    new_height = height - (height % block_size)
    new_width = width - (width % block_size)
    cropped_img = img[:new_height, :new_width]  # 裁剪多余像素
    return cropped_img


def generate_heatmap(image, block_size):
    h, w = image.shape
    heatmap = np.zeros((h // block_size, w // block_size))
    for i in range(0, h, block_size):
        for j in range(0, w, block_size):
            block = image[i: i + block_size, j: j + block_size]
            heatmap[i // block_size, j //
                    block_size] = np.mean(block)  # 计算块的平均值
    return heatmap


def capture_dark_heatmap(
    colorimeter: mlcm.ML_Colorimeter,
    binn_selector:mlcm.BinningSelector,
    binn_mode:mlcm.BinningMode,
    binn:mlcm.Binning,
    pixel_format:mlcm.MLPixelFormat,
    nd_list: List[mlcm.MLFilterEnum],
    xyz_list: List[mlcm.MLFilterEnum],
    binn_list: List[mlcm.Binning],
    et_list: List[float],
    save_path: str,
    file_name: str,
    capture_times: int,
    status_callback=None
):
    def update_status(message):
        if status_callback:
            status_callback(message)
    # update_status("capture_dark_heatmap start...")
    # time.sleep(10)
    # update_status("capture_dark_heatmap finish...")

    module_id = 1
    mono = colorimeter.ml_bino_manage.ml_get_module_by_id(module_id)

    ret = mono.ml_set_binning_selector(binn_selector)
    if not ret.success:
        raise RuntimeError("ml_set_binning_selector error")

    # Set binning mode for camera.
    ret = mono.ml_set_binning_mode(binn_mode)
    if not ret.success:
        raise RuntimeError("ml_set_binning_mode error")

    # Format of the pixel to use for acquisition.
    ret = mono.ml_set_pixel_format(pixel_format)
    if not ret.success:
        raise RuntimeError("ml_set_pixel_format error")

    for nd in nd_list:
        nd_enum = mlcm.MLFilterEnum(nd)
        ret = mono.ml_move_nd_syn(nd_enum)
        if not ret.success:
            raise RuntimeError("ml_move_nd_syn error")

        for xyz in xyz_list:
            xyz_enum = mlcm.MLFilterEnum(xyz)
            ret = mono.ml_move_xyz_syn(xyz_enum)
            if not ret.success:
                raise RuntimeError("ml_move_xyz_syn error")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            temp_str = mlcm.MLFilterEnum_to_str(
                nd_enum) + "_" + mlcm.MLFilterEnum_to_str(xyz_enum) + "_" + timestamp +"_"
            file_path = save_path + "\\" + temp_str + file_name
            wb = Workbook()
            wb.save(file_path)
            wb = load_workbook(file_path)
            wb.remove(wb["Sheet"])
            for binn in binn_list:
                ret = mono.ml_set_binning(mlcm.Binning(binn))
                if not ret.success:
                    raise RuntimeError("ml_set_binning error")
                get_binn = mono.ml_get_binning()
                # print(get_binn)

                # generate xlsx file
                title = str(pow(2, binn)) + "X" + str(pow(2, binn))
                ws = wb.create_sheet(title=title)

                gray_list = []
                raw_list = []
                raw_max_list = []
                processed_max_list = []
                img_idx = 19
                for et in et_list:
                    exposure = mlcm.pyExposureSetting(
                        exposure_mode=mlcm.ExposureMode.Fixed, exposure_time=et
                    )

                    # Set exposure for camera, contain auto and fixed.
                    ret = mono.ml_set_exposure(exposure)
                    if not ret.success:
                        raise RuntimeError("ml_set_exposure error")

                    # capture a single image
                    list = []
                    for i in range(capture_times):
                        ret = mono.ml_capture_image_syn()
                        if not ret.success:
                            raise RuntimeError("ml_capture_image_syn error")
                        get_img = mono.ml_get_image()
                        list.append(get_img)
                    average_image = np.mean(list, axis=0).astype(np.uint16)
                    gray_list.append(cv2.mean(average_image)[0])
                    raw_list.append(average_image)
                    if not os.path.exists(save_path + "\\raw"):
                        os.makedirs(save_path + "\\raw")
                    cv2.imwrite(save_path + "\\raw\\" + str(pow(2, binn)) + "X" +
                                str(pow(2, binn)) + "_" + str(et) + "ms.tif", average_image)

                for k in range(et_list.__len__()):
                    exposure = mlcm.pyExposureSetting(
                        exposure_mode=mlcm.ExposureMode.Fixed, exposure_time=et_list[k]
                    )

                    # Set exposure for camera, contain auto and fixed.
                    ret = mono.ml_set_exposure(exposure)
                    if not ret.success:
                        raise RuntimeError("ml_set_exposure error")

                    # capture a single image
                    ret = mono.ml_capture_image_syn()
                    if not ret.success:
                        raise RuntimeError("ml_capture_image_syn error")
                    get_img = mono.ml_get_image()

                    block_size = 10
                    cropped_img = preprocess_image(get_img, block_size)
                    heatmap = generate_heatmap(cropped_img, block_size)
                    # plt.imshow(heatmap, cmap="jet", interpolation="nearest")
                    plt.imshow(heatmap, cmap="jet", interpolation="nearest", vmin=0, vmax=10)
                    plt.colorbar()
                    plt.title("raw_" + str(pow(2, binn)) + "X" +
                            str(pow(2, binn)) + "_" + str(et_list[k]) + "ms")
                    plt.savefig(save_path + "\\raw\\raw_" + str(pow(2, binn)) +
                                "X" + str(pow(2, binn)) + "_" + str(et_list[k]) + "ms.png")
                    # plt.show()
                    plt.close()
                    raw_max_list.append(np.max(heatmap))

                    img = Image(save_path + "\\raw\\raw_" + str(pow(2, binn)) +
                                "X" + str(pow(2, binn)) + "_" + str(et_list[k]) + "ms.png")

                    img.width = 300  # 像素宽度
                    img.height = 200  # 像素高度
                    ws[f"B{int(img_idx + 15*k)}"] = "raw " + \
                        str(et_list[k]) + " ms"
                    ws.add_image(img, f"B{int(img_idx + 2 + 15*k)}")

                    if not os.path.exists(save_path + "\\processed"):
                        os.makedirs(save_path + "\\processed")
                    get_img = np.subtract(np.array(get_img, dtype=np.int16), np.array(
                        raw_list[k], dtype=np.int16))
                    cropped_img = preprocess_image(get_img, block_size)
                    heatmap = generate_heatmap(cropped_img, block_size)
                    plt.imshow(heatmap, cmap="jet", interpolation="nearest", vmin=0, vmax=10)
                    # plt.imshow(heatmap, cmap="jet", interpolation="nearest")
                    plt.colorbar()
                    plt.title("processed_" + str(pow(2, binn)) + "X" +
                            str(pow(2, binn)) + "_" + str(et_list[k]) + "ms")
                    plt.savefig(save_path + "\\processed\\processed_" + str(pow(2, binn)) +
                                "X" + str(pow(2, binn)) + "_" + str(et_list[k]) + "ms.png")
                    plt.close()
                    img2 = Image(save_path + "\\processed\\processed_" + str(pow(2, binn)) +
                                "X" + str(pow(2, binn)) + "_" + str(et_list[k]) + "ms.png")

                    img2.width = 300  # 像素宽度
                    img2.height = 200  # 像素高度
                    ws[f"I{int(img_idx + 15*k)}"] = "processed " + \
                        str(et_list[k]) + " ms"
                    ws.add_image(img2, f"I{int(img_idx + 2 + 15*k)}")

                    arr32 = heatmap.astype(np.float32)
                    cv2.imwrite(save_path + "\\processed\\" +
                                str(et_list[k]) + "ms.tif", arr32)
                    processed_max_list.append(np.max(arr32))

                ws["B1"] = "No."
                ws["C1"] = "ET(ms)"
                ws["D1"] = "Grayscale"
                ws["E1"] = "10X10 pixel raw max"
                ws["F1"] = "10X10 pixel processed max"

                for row in range(et_list.__len__()):
                    ws[f"B{row+2}"] = row + 1
                    ws[f"C{row+2}"] = et_list[row]
                    ws[f"D{row+2}"] = gray_list[row]
                    ws[f"E{row+2}"] = raw_max_list[row]
                    ws[f"F{row+2}"] = processed_max_list[row]

                wb.save(file_path)

            update_status("finish")


if __name__ == "__main__":
    # set mono module calibration configuration path
    eye1_path = r"D:\MLOptic\MLColorimeter\config\EYE1"
    save_path = r"E:\project\test"
    if not os.path.exists(save_path):
        os.makedirs(save_path)

    path_list = [
        eye1_path,
    ]
    try:
        # create a ML_Colorimeter system instance
        ml_colorimeter = mlcm.ML_Colorimeter()
        # add mono module into ml_colorimeter system, according to path_list create one or more mono module
        ret = ml_colorimeter.ml_add_module(path_list=path_list)
        if not ret.success:
            raise RuntimeError("ml_add_module error")
        # connect all module in the ml_colorimeter system
        ret = ml_colorimeter.ml_connect()
        if not ret.success:
            raise RuntimeError("ml_connect error")

        # 4:ND0  5:ND1  6:ND2  7:ND3  8:ND4
        nd_list = [4]
        # 1:X  2:Y  3:Z  10:Clear
        xyz_list = [1]
        # exposure time list
        et_list = [0.1, 1, 10, 20, 50]
        # 0:ONE_BY_ONE  1:TWO_BY_TWO  2:FOUR_BY_FOUR  3:EIGHT_BY_EIGHT  4:SIXTEEN_BY_SIXTEEN
        binn_list = [0]
        # file name
        file_name = "dark_heatmap.xlsx"
        # multi frame averaging
        capture_times = 5
        capture_dark_heatmap(ml_colorimeter, nd_list, xyz_list,
                            binn_list, et_list, save_path, file_name, capture_times)
    except Exception as e:
        # print(e)
        pass
