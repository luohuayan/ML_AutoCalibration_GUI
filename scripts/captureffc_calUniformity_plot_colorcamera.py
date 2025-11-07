import mlcolorimeter as mlcm
from typing import List, Dict
import cv2
import csv
import os
import numpy as np
from datetime import datetime
from threading import Thread
import time
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook


__all__ = ["cal_synthetic_mean_images", "capture_ffc_images", "cal_uniformity"]


def datetime_str():
    return datetime.now().strftime('%Y-%m-%d_%H-%M-%S.%f')[:-3]


def datetime_str2():
    return datetime.now().strftime('%Y-%m-%d_%H-%M-%S')


def measurement2(
    colorimeter: mlcm.ML_Colorimeter,
    half_size:int,
    vrange: List,
    nd: mlcm.MLFilterEnum,
    xyz_list: List[mlcm.MLFilterEnum],
    RX_str: str,
    binn: mlcm.Binning,
    exposure: mlcm.pyExposureSetting,
    module_id: int,
    cali_config: mlcm.pyCalibrationConfig,
    save_config: mlcm.pySaveDataConfig,
    roi_dict: Dict[mlcm.Binning, List[mlcm.pyCVRect]],
    ffc_ws,
    fourcolor_ws,
    ffc_wb,
    fourcolor_wb,
    ffc_xlsx,
    fourcolor_xlsx
):
    mono = colorimeter.ml_bino_manage.ml_get_module_by_id(module_id)
    # capture data
    capture_data_dict = dict()
    ret = ml_mono.ml_set_exposure(exposure)
    if not ret.success:
        raise RuntimeError("ml_set_exposure error")

    ret = ml_mono.ml_capture_image_syn()
    if not ret.success:
        raise RuntimeError("ml_capture_image_syn error")

    capture_data_dict = ml_mono.ml_get_colorcamera_CaptureData()
    # set capture data for measurement
    ret = colorimeter.ml_set_CaptureData(module_id, capture_data_dict)
    if not ret.success:
        raise RuntimeError("ml_set_CaptureData error")

    # load calibration data by calbration config
    ret = colorimeter.ml_load_calibration_data(cali_config)
    if not ret.success:
        raise RuntimeError("ml_load_calibration_data error")

    # execute calibration process for capture data by calibration config
    ret = colorimeter.ml_image_process(cali_config)
    if not ret.success:
        raise RuntimeError("ml_image_process error")

    # get calibration data after calibration process
    processed_data = colorimeter.ml_get_processed_data(module_id)

    # save calibration data
    ret = colorimeter.ml_save_processed_data(
        module_id, processed_data, save_config)
    if not ret.success:
        raise RuntimeError("ml_save_processed_data error")

    roi_list = roi_dict[binn]

    for xyz in xyz_list:
        img = processed_data[mlcm.CalibrationEnum.FFC][xyz].image
        line = []
        color_filter = mlcm.MLFilterEnum_to_str(xyz)
        nd_filter = mlcm.MLFilterEnum_to_str(nd)
        line.append(RX_str)
        line.append(color_filter)
        line.append(nd_filter)
        line.append(ml_mono.ml_get_light_source())

        gray_list = []
        for rect in roi_list:
            roi_img = img[rect.y: rect.y+rect.height,
                          rect.x: rect.x+rect.width]
            gray = cv2.mean(roi_img)[0]
            gray_list.append(gray)
            line.append(gray)
        line.append(np.mean(gray_list))
        line.append(np.std(gray_list, ddof=1))
        line.append(min(gray_list))
        line.append(max(gray_list))
        line.append(min(gray_list) / max(gray_list))

        uniformity = 1 - 0.5 * \
            (max(gray_list) - min(gray_list)) / np.mean(gray_list)
        line.append(uniformity)

        ffc_ws.append(line)
        line = []
        gray_list = []

        row, col = img.shape[0], img.shape[1]
        center_row, center_col = int(row / 2), int(col / 2)
        half_size = half_size / mlcm.Binning_to_Int(binn)

        fig = plt.figure(figsize=(16, 4))
        plt.subplot(1, 2, 1)
        plt.imshow(img, cmap="jet", vmin=vrange[0], vmax=vrange[1])
        plt.xlabel("Col(pixel)")
        plt.ylabel("Row(pixel)")
        plt.colorbar(label="(FFC)")
        plt.title("FFC_" + mlcm.Binning_to_str(binn) +
                  "_" + mlcm.MLFilterEnum_to_str(xyz))
        plt.subplot(1, 2, 2)
        plt.plot(
            img[center_row, int(center_col - half_size)
                                : int(center_col + half_size)]
        )
        plt.plot(
            img[int(center_row - half_size)
                    : int(center_row + half_size), center_col]
        )
        plt.grid(which="both")
        plt.xlabel("Position(pixel)")
        plt.ylabel("gray value")
        plt.ylim(vrange)
        plt.legend(["Along-Col", "Along-Row"])
        # 显示图形
        # plt.show(block=True)
        # 保存图形
        plt.savefig(
            uniformity_path + "\\FFC_ " +
            mlcm.Binning_to_str(binn) + "_" + mlcm.MLFilterEnum_to_str(xyz) + ".png",
            bbox_inches="tight",
        )
        plt.close(fig)

    ffc_wb.save(ffc_xlsx)

    fourcolor_img_list = []
    fourcolor_img_name_list = ["Luminance", "Cx", "Cy"]
    luminance_img = processed_data[mlcm.CalibrationEnum.FinalResult][mlcm.MLFilterEnum.Y].image
    fourcolor_img_list.append(luminance_img)

    for xyz in [mlcm.MLFilterEnum.X, mlcm.MLFilterEnum.Y]:
        img = processed_data[mlcm.CalibrationEnum.Chrom][xyz].image
        fourcolor_img_list.append(img)

    for i in range(len(fourcolor_img_list)):
        line = []
        color_filter = mlcm.MLFilterEnum_to_str(xyz)
        nd_filter = mlcm.MLFilterEnum_to_str(nd)
        line.append(RX_str)
        line.append(fourcolor_img_name_list[i])
        line.append(nd_filter)
        line.append(ml_mono.ml_get_light_source())

        gray_list = []
        for rect in roi_list:
            roi_img = fourcolor_img_list[i][rect.y: rect.y +
                                            rect.height, rect.x: rect.x+rect.width]
            gray = cv2.mean(roi_img)[0]
            gray_list.append(gray)
            line.append(gray)

        line.append(np.mean(gray_list))
        line.append(np.std(gray_list, ddof=1))
        line.append(min(gray_list))
        line.append(max(gray_list))
        line.append(min(gray_list) / max(gray_list))
        fourcolor_ws.append(line)
        line = []

    fourcolor_wb.save(fourcolor_xlsx)


def cal_synthetic_mean_images2(
    colorimeter: mlcm.ML_Colorimeter,
    nd_list: List[mlcm.MLFilterEnum],
    xyz_list: List[mlcm.MLFilterEnum],
    save_path: str,
):
    module_id = 1
    mono = colorimeter.ml_bino_manage.ml_get_module_by_id(module_id)
    for nd in nd_list:
        for xyz in xyz_list:
            colorimeter.ml_cal_synthetic_mean_images(
                module_id=module_id,
                save_path=save_path,
                aperture=mono.ml_get_aperture(),
                nd=nd,
                xyz=xyz,
                sphere_list=[0],
                light_source=mono.ml_get_light_source()
            )
            print("calculate mean images for: " +
                  mlcm.MLFilterEnum_to_str(xyz))
        print("calculate mean images for: " +
              mlcm.MLFilterEnum_to_str(nd))
    print("calculate ffc synthetic mean finish")


def capture_ffc_images2(
    colorimeter: mlcm.ML_Colorimeter,
    nd_list: List[mlcm.MLFilterEnum],
    binn: int,
    exposure_map: Dict[mlcm.MLFilterEnum, mlcm.pyExposureSetting],
    capture_times: int,
    save_path: str,
    use_RX: bool = False,
    sph_list: List = [],
    cyl_list: List = [],
    axis_list: List = []
):
    module_id = 1
    mono = colorimeter.ml_bino_manage.ml_get_module_by_id(module_id)

    for nd in nd_list:
        ret = mono.ml_move_nd_syn(nd)
        print(mlcm.MLFilterEnum_to_str(nd))
        if not ret.success:
            raise RuntimeError("ml_move_nd_syn error")

        ret = mono.ml_set_binning(binn)
        if not ret.success:
            raise RuntimeError("ml_set_binning error")
        get_binn = mono.ml_get_binning()
        print(mlcm.Binning_to_str(get_binn))

        if use_RX == False or (not sph_list) or (not cyl_list) or (not axis_list):
            rx = mlcm.pyRXCombination(0, 0, 0)
            # capture ffc images
            ret = colorimeter.ml_capture_ffc_colorcamera2(
                module_id=module_id,
                save_path=save_path,
                nd=nd,
                rx=rx,
                avg_count=capture_times,
                exposure_map=exposure_map[nd]
            )
            if not ret.success:
                raise RuntimeError("ml_capture_ffc_colorcamera error")

        else:
            for sph in sph_list:
                rx = mlcm.pyRXCombination(sph, 0, 0)
                ret = colorimeter.ml_capture_ffc_colorcamera2(
                    module_id=module_id,
                    save_path=save_path,
                    nd=nd,
                    rx=rx,
                    avg_count=capture_times,
                    exposure_map=exposure_map[nd]
                )
                print("capture image for: " + mlcm.pyRXCombination_to_str(rx))
                if not ret.success:
                    raise RuntimeError("ml_capture_ffc_colorcamera error")

            for cyl in cyl_list:
                for axis in axis_list:
                    rx = mlcm.pyRXCombination(0, cyl, axis)
                    ret = colorimeter.ml_capture_ffc_colorcamera(
                        module_id=module_id,
                        save_path=save_path,
                        nd=nd,
                        rx=rx,
                        avg_count=capture_times,
                        exposure=exposure_map[nd]
                    )
                    print("capture image for: " +
                          mlcm.pyRXCombination_to_str(rx))
                    if not ret.success:
                        raise RuntimeError("ml_capture_ffc_colorcamera error")

    print("capture color camera ffc images finish")


def cal_uniformity2(
    colorimeter: mlcm.ML_Colorimeter,
    half_size: int,
    vrange: List,
    nd_list: List[mlcm.MLFilterEnum],
    xyz_list: List[mlcm.MLFilterEnum],
    uniformity_path: str,
    binn_list: List[mlcm.Binning],
    exposure: mlcm.pyExposureSetting,
    roi_dict: Dict[mlcm.Binning, List[mlcm.pyCVRect]],
    use_RX: bool = False,
    rx_dict: Dict[mlcm.MLFilterEnum, List] = None,
):
    module_id = 1
    mono = colorimeter.ml_bino_manage.ml_get_module_by_id(module_id)

    cali_path = mono.ml_get_config_path()

    for nd in nd_list:
        ffc_xlsx = uniformity_path + r"\ffc_uniformity-" + datetime_str2() + "-" + \
            mlcm.MLFilterEnum_to_str(nd) + "_" + \
            mono.ml_get_aperture()+".xlsx"

        fourcolor_xlsx = uniformity_path + r"\fourcolor_uniformity-" + datetime_str2() + "-" + \
            mlcm.MLFilterEnum_to_str(nd) + "_" + \
            mono.ml_get_aperture()+".xlsx"

        ret = ml_mono.ml_move_nd_syn(nd)
        if not ret.success:
            raise RuntimeError("ml_move_nd_syn error")

        if use_RX:
            sph_list = rx_dict[nd][0]
            cyl_list = rx_dict[nd][1]
            axis_list = rx_dict[nd][2]

        for binn in binn_list:
            ret = mono.ml_set_binning(binn)
            if not ret.success:
                raise RuntimeError("ml_set_binning error")
            get_binn = mono.ml_get_binning()
            print(get_binn)

            ffc_wb = Workbook()
            ffc_wb.save(ffc_xlsx)
            ffc_wb = load_workbook(ffc_xlsx)
            ffc_wb.remove(ffc_wb["Sheet"])
            title = str(pow(2, int(binn))) + "X" + str(pow(2, int(binn)))
            ffc_ws = ffc_wb.create_sheet(title=title)
            ffc_title = ["RX", "ColorFilter", "NDFilter", "Light Source", "ROI1", "ROI2", "ROI3",
                         "ROI4", "ROI5", "ROI6", "ROI7", "ROI8", "ROI9", "Mean", "Std", "Min", "Max", "Min/Max", "Uniformity"]
            ffc_ws.append(ffc_title)

            fourcolor_wb = Workbook()
            fourcolor_wb.save(fourcolor_xlsx)
            fourcolor_wb = load_workbook(fourcolor_xlsx)
            fourcolor_wb.remove(fourcolor_wb["Sheet"])
            title = str(pow(2, int(binn))) + "X" + str(pow(2, int(binn)))
            fourcolor_ws = fourcolor_wb.create_sheet(title=title)
            fourcolor_title = ["RX", "CxCyLuminance", "NDFilter", "Light Source", "ROI1", "ROI2", "ROI3",
                                     "ROI4", "ROI5", "ROI6", "ROI7", "ROI8", "ROI9", "Mean", "Std", "Min", "Max", "Min/Max", "Uniformity"]
            fourcolor_ws.append(fourcolor_title)

            if not use_RX:
                rx = mlcm.pyRXCombination(0, 0, 0)
                RX_str = mlcm.pyRXCombination_to_str(rx)

                # calibration config for measurement
                cali_config = mlcm.pyCalibrationConfig(
                    input_path=cali_path, 
                    aperture=mono.ml_get_aperture(),
                    nd_filter_list=[nd], 
                    color_filter_list=xyz_list,
                    rx=rx, 
                    light_source_list=[mono.ml_get_light_source()],
                    dark_flag=True,
                    ffc_flag=True,
                    color_shift_flag=True,
                    distortion_flag=True,
                    exposure_flag=True,
                    four_color_flag=True,
                )

                # save config
                save_config = mlcm.pySaveDataConfig(
                    save_path=uniformity_path + f"\\res",
                    save_raw=False,
                    save_result=True,
                    save_calibration=False
                )

                measurement2(
                    colorimeter, 
                    half_size,
                    vrange,
                    nd, 
                    xyz_list, 
                    RX_str, 
                    binn, 
                    exposure, 
                    module_id, 
                    cali_config, 
                    save_config,
                    roi_dict, 
                    ffc_ws, 
                    fourcolor_ws, 
                    ffc_wb, 
                    fourcolor_wb, 
                    ffc_xlsx, 
                    fourcolor_xlsx
                )

            else:
                for sph in sph_list:
                    for cyl in cyl_list:
                        for axis in axis_list:
                            rx = mlcm.pyRXCombination(sph, cyl, axis)
                            ret = ml_mono.ml_set_rx_syn(rx)
                            if not ret.success:
                                raise RuntimeError("ml_set_rx_syn error")

                            RX_str = mlcm.pyRXCombination_to_str(rx)

                            # calibration config for measurement
                            cali_config = mlcm.pyCalibrationConfig(
                                input_path=cali_path, 
                                aperture=mono.ml_get_aperture(), 
                                nd_filter_list=[nd],
                                color_filter_list=xyz_list, 
                                rx=rx, 
                                light_source_list=[mono.ml_get_light_source()],
                                dark_flag=True,
                                ffc_flag=True,
                                color_shift_flag=True,
                                distortion_flag=True,
                                exposure_flag=True,
                                four_color_flag=True,
                            )

                            # save config
                            save_config = mlcm.pySaveDataConfig(
                                save_path=uniformity_path + f"\\res",
                                save_raw=False,
                                save_result=True,
                                save_calibration=False
                            )

                            measurement2(
                                colorimeter, 
                                half_size,
                                vrange,
                                nd, 
                                xyz_list, 
                                RX_str, 
                                binn, 
                                exposure, 
                                module_id, 
                                cali_config,
                                save_config, 
                                roi_dict, 
                                ffc_ws, 
                                fourcolor_ws, 
                                ffc_wb, 
                                fourcolor_wb, 
                                ffc_xlsx, 
                                fourcolor_xlsx
                            )
                            print("measurement for rx: " + RX_str)


if __name__ == '__main__':
    eye1_path = r"G:\project\weilaixing\EYE1"
    path_list = [
        eye1_path,
    ]
    uniformity_path = r"G:\aolanduo2_uniformity"
    if not os.path.exists(uniformity_path):
        os.makedirs(uniformity_path)
    try:
        # create a ML_Colorimeter system instance
        ml_colorimeter = mlcm.ML_Colorimeter()
        # add mono module into ml_colorimeter system, according to path_list create one or more mono module
        ret = ml_colorimeter.ml_add_module(path_list=path_list)
        if not ret.success:
            raise RuntimeError("ml_add_module error")
        # connect all module in the ml_colorimeter system
        ret = ml_colorimeter.ml_connect()
        if not ret.success:
            raise RuntimeError("ml_connect error")

        module_id = 1
        ml_mono = ml_colorimeter.ml_bino_manage.ml_get_module_by_id(module_id)

        light_source = "W"
        ml_mono.ml_set_light_source(light_source)

        pixel_format = mlcm.MLPixelFormat.MLBayerRG12
        ret = ml_mono.ml_set_pixel_format(pixel_format)
        if not ret.success:
            raise RuntimeError("ml_set_pixel_format error")

        # 4:ND0  5:ND1  6:ND2  7:ND3  8:ND4
        nd_list = [4]
        nd_list = [mlcm.MLFilterEnum(int(nd)) for nd in nd_list]
        # 1:X  2:Y  3:Z  10:Clear
        xyz_list = [1, 2, 3]
        xyz_list = [mlcm.MLFilterEnum(int(xyz)) for xyz in xyz_list]
        # multi frame averaging
        capture_times = 1
        # different exposure map of nd while capture ffc images
        exposure_map_obj = {
            mlcm.MLFilterEnum.ND0: {
                mlcm.MLFilterEnum.X: mlcm.pyExposureSetting(exposure_mode=mlcm.ExposureMode.Fixed, exposure_time=200),
                mlcm.MLFilterEnum.Y: mlcm.pyExposureSetting(exposure_mode=mlcm.ExposureMode.Fixed, exposure_time=100),
                mlcm.MLFilterEnum.Z: mlcm.pyExposureSetting(exposure_mode=mlcm.ExposureMode.Fixed, exposure_time=200),
            },
            mlcm.MLFilterEnum.ND1: {
                mlcm.MLFilterEnum.X: mlcm.pyExposureSetting(exposure_mode=mlcm.ExposureMode.Fixed, exposure_time=120),
                mlcm.MLFilterEnum.Y: mlcm.pyExposureSetting(exposure_mode=mlcm.ExposureMode.Fixed, exposure_time=90),
                mlcm.MLFilterEnum.Z: mlcm.pyExposureSetting(exposure_mode=mlcm.ExposureMode.Fixed, exposure_time=280),
            },
            mlcm.MLFilterEnum.ND2: {
                mlcm.MLFilterEnum.X: mlcm.pyExposureSetting(exposure_mode=mlcm.ExposureMode.Fixed, exposure_time=4556),
                mlcm.MLFilterEnum.Y: mlcm.pyExposureSetting(exposure_mode=mlcm.ExposureMode.Fixed, exposure_time=4054),
                mlcm.MLFilterEnum.Z: mlcm.pyExposureSetting(exposure_mode=mlcm.ExposureMode.Fixed, exposure_time=4999),
            },
            mlcm.MLFilterEnum.ND3: {
                mlcm.MLFilterEnum.X: mlcm.pyExposureSetting(exposure_mode=mlcm.ExposureMode.Fixed, exposure_time=4556),
                mlcm.MLFilterEnum.Y: mlcm.pyExposureSetting(exposure_mode=mlcm.ExposureMode.Fixed, exposure_time=4054),
                mlcm.MLFilterEnum.Z: mlcm.pyExposureSetting(exposure_mode=mlcm.ExposureMode.Fixed, exposure_time=4999),
            },
            mlcm.MLFilterEnum.ND4: {
                mlcm.MLFilterEnum.X: mlcm.pyExposureSetting(exposure_mode=mlcm.ExposureMode.Fixed, exposure_time=4556),
                mlcm.MLFilterEnum.Y: mlcm.pyExposureSetting(exposure_mode=mlcm.ExposureMode.Fixed, exposure_time=4054),
                mlcm.MLFilterEnum.Z: mlcm.pyExposureSetting(exposure_mode=mlcm.ExposureMode.Fixed, exposure_time=4999),
            }
        }

        # 0:ONE_BY_ONE  1:TWO_BY_TWO  2:FOUR_BY_FOUR  3:EIGHT_BY_EIGHT  4:SIXTEEN_BY_SIXTEEN
        # binn of capture ffc images
        binn = 0
        binn = mlcm.Binning(int(binn))
        # binn list of calculate uniformity
        binn_list = [0]
        binn_list = [mlcm.Binning(int(binn)) for binn in binn_list]
        # exposure for calculate uniformity
        exposure = mlcm.pyExposureSetting(
            exposure_mode=mlcm.ExposureMode.Auto, exposure_time=100)

        use_RX = False
        # sph, cyl, axis list while capture ffc images
        sph_list = [-6, -5, -4, -3, -2, -1, 0, 1, 2, 3, 4, 5, 6]
        cyl_list = [-4, -3.5, -3, -2.5, -2, -1.5, -1, -0.5, 0]
        axis_list = [0, 15, 30, 45, 60, 75, 90, 105, 120, 135, 150, 165]
        save_path = eye1_path

        half_size = 3600
        vrange = [15, 390]

        # different rx list of nd while calculate uniformity
        rx_dict = {
            mlcm.MLFilterEnum.ND0: [
                [-6, -5, -4, -3, -2, -1, 0, 1, 2, 3, 4],
                [-4, -3.5, -3, -2.5, -2, -1.5, -1, -0.5, 0],
                [0, 90]
            ],
            mlcm.MLFilterEnum.ND1: [
                [-6, -5, -4, -3, -2, -1, 0, 1, 2, 3, 4],
                [-4, -3.5, -3, -2.5, -2, -1.5, -1, -0.5, 0],
                [0, 90]
            ],
            mlcm.MLFilterEnum.ND2: [
                [-6, -5, -4, -3, -2, -1, 0, 1, 2, 3, 4],
                [-4, -3.5, -3, -2.5, -2, -1.5, -1, -0.5, 0],
                [0, 90]
            ],
            mlcm.MLFilterEnum.ND3: [
                [-6, -5, -4, -3, -2, -1, 0, 1, 2, 3, 4],
                [-4, -3.5, -3, -2.5, -2, -1.5, -1, -0.5, 0],
                [0, 90]
            ],

        }

        # different roi list of binn while calculate uniformity
        roi_dict = {
            mlcm.Binning.ONE_BY_ONE: [
                mlcm.pyCVRect(3000, 2632, 400, 400),
                mlcm.pyCVRect(3000, 4632, 400, 400),
                mlcm.pyCVRect(3000, 6632, 400, 400),
                mlcm.pyCVRect(4900, 2632, 400, 400),
                mlcm.pyCVRect(4900, 4632, 400, 400),
                mlcm.pyCVRect(4900, 6632, 400, 400),
                mlcm.pyCVRect(7800, 2632, 400, 400),
                mlcm.pyCVRect(7800, 4632, 400, 400),
                mlcm.pyCVRect(7800, 6632, 400, 400),
            ],
            mlcm.Binning.TWO_BY_TWO: [
                mlcm.pyCVRect(3000, 2632, 200, 200),
                mlcm.pyCVRect(3000, 4632, 200, 200),
                mlcm.pyCVRect(3000, 6632, 200, 200),
                mlcm.pyCVRect(4900, 2632, 200, 200),
                mlcm.pyCVRect(4900, 4632, 200, 200),
                mlcm.pyCVRect(4900, 6632, 200, 200),
                mlcm.pyCVRect(7800, 2632, 200, 200),
                mlcm.pyCVRect(7800, 4632, 200, 200),
                mlcm.pyCVRect(7800, 6632, 200, 200),
            ],
            mlcm.Binning.FOUR_BY_FOUR: [
                mlcm.pyCVRect(800, 658, 100, 100),
                mlcm.pyCVRect(800, 1158, 100, 100),
                mlcm.pyCVRect(800, 1658, 100, 100),
                mlcm.pyCVRect(1350, 658, 100, 100),
                mlcm.pyCVRect(1350, 1158, 100, 100),
                mlcm.pyCVRect(1350, 1658, 100, 100),
                mlcm.pyCVRect(1900, 658, 100, 100),
                mlcm.pyCVRect(1900, 1158, 100, 100),
                mlcm.pyCVRect(1900, 1658, 100, 100),
            ]
        }

        is_capture_ffc = True
        is_calculate_synthetic = True
        is_calculate_uniformity = True

        if is_capture_ffc:
            capture_ffc_images2(
                ml_colorimeter,
                nd_list,
                binn,
                exposure_map_obj,
                capture_times,
                save_path,
                use_RX,
                sph_list,
                cyl_list,
                axis_list
            )

        if use_RX and is_calculate_synthetic:
            cal_synthetic_mean_images2(
                ml_colorimeter,
                nd_list,
                xyz_list,
                save_path
            )

        if is_calculate_uniformity:
            cal_uniformity2(
                ml_colorimeter,
                half_size,
                vrange,
                nd_list,
                xyz_list,
                uniformity_path,
                binn_list,
                exposure,
                roi_dict,
                use_RX,
                rx_dict
            )

    except Exception as e:
        print(e)
